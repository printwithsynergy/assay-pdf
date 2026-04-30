"""Parse the GWG 2022 specification workbook into a typed RequirementManifest.

Workbook structure (as of 2026-04 fetch):
- 'Legend' — explanatory text (skipped)
- 'Definitions' — D-prefixed terms (skipped for v0.1.0; available for future enrichment)
- 'Requirements' — R-prefixed rules with title, text, notes, discussion
- 'Variants' — per-rule by per-variant severity matrix
- 'Processing Steps' — ISO 19593-1 layer-type rules (modeled separately, future commit)
- 'Product Types' — D0029 product type definitions (skipped)
- 'Implementation Notes' — vendor guidance (skipped)

Severity cells in the Variants tab take one of four values:
- "error" / "warning" / "ignore" / "error\\nwarning" (dual-band)
"ignore" means "rule does not apply" — those variants are omitted from the
applicability dict (see Severity docstring for rationale).
"""

from __future__ import annotations

import re
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

from assay_pdf import __version__
from assay_pdf.hashing import sha256_file
from assay_pdf.logging import get_logger
from assay_pdf.models import (
    Requirement,
    RequirementManifest,
    RuleApplicability,
    Severity,
    Variant,
)

log = get_logger(__name__)

_VARIANT_SHEET = "Variants"
_REQS_SHEET = "Requirements"

# Spec literal severity strings → Severity enum mapping.
# Both ordering of "error\nwarning" and "warning\nerror" map to error_and_warning.
_SEVERITY_LITERALS: dict[str, Severity] = {
    "error": Severity.error,
    "warning": Severity.warning,
    "error\nwarning": Severity.error_and_warning,
    "warning\nerror": Severity.error_and_warning,
}


def _kebab(name: str) -> str:
    """Normalize a variant name to filesystem-safe kebab-case.

    'Magazine Ads CMYK + RGB' -> 'magazine-ads-cmyk-plus-rgb'
    'Folding Carton & Corrugated Box' -> 'folding-carton-corrugated-box'
    """
    s = re.sub(r"\+", "plus", name)
    s = re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").lower()
    return s


def _parse_additional_values(raw: object) -> dict[str, str]:
    """Parse 'A = 224ppi, B = 5mm' style cells into a {key: value} dict."""
    if raw is None:
        return {}
    text = str(raw).replace("\n", ",")
    out: dict[str, str] = {}
    for part in text.split(","):
        if "=" not in part:
            continue
        k, v = part.split("=", 1)
        out[k.strip()] = v.strip()
    return out


def _coerce_severity(raw: object, *, rule_id: str, variant: str) -> Severity | None:
    """Map a spec severity cell to a Severity enum value.

    Returns None when:
    - The cell is empty (rule simply not in this row's applicability)
    - The cell is "ignore" (rule explicitly does not apply — also returns None)

    Logs a warning and returns None for any other unexpected value.
    """
    if raw is None:
        return None
    value = str(raw).strip().lower()
    if value == "" or value == "ignore":
        return None
    severity = _SEVERITY_LITERALS.get(value)
    if severity is None:
        log.warning(
            "Unrecognized severity %r for rule %s variant %r — treating as not-applicable",
            value,
            rule_id,
            variant,
        )
    return severity


def _coerce_version(raw: object) -> float:
    """Best-effort float conversion for the Version column."""
    if raw is None:
        return 1.0
    try:
        return float(str(raw))
    except (TypeError, ValueError):
        return 1.0


def parse_workbook(xlsx_path: Path | str) -> RequirementManifest:
    """Parse the GWG 2022 XLSX into a typed RequirementManifest."""
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Spec workbook not found at {path}")

    log.info("Parsing GWG 2022 workbook: %s", path)
    wb = openpyxl.load_workbook(path, data_only=True)

    if _VARIANT_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook missing required tab {_VARIANT_SHEET!r}; got {wb.sheetnames}")
    if _REQS_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook missing required tab {_REQS_SHEET!r}; got {wb.sheetnames}")

    # ─── Variants tab: extract variant column positions and per-rule severity ───
    ws_var = wb[_VARIANT_SHEET]
    variant_cols: dict[str, int] = {}  # variant_name → 1-based openpyxl column
    variants: list[Variant] = []
    for cell in ws_var[1]:
        if cell.value:
            name = str(cell.value).strip()
            variant_cols[name] = cell.column
            variants.append(Variant(name=name, kebab=_kebab(name)))

    if not variants:
        raise ValueError(f"No variant headers found in row 1 of {_VARIANT_SHEET!r}")

    log.info("Found %d variants", len(variants))

    # Per-rule applicability map
    applicability: dict[str, dict[str, RuleApplicability]] = {}
    for row_idx in range(3, ws_var.max_row + 1):
        rid_value = ws_var.cell(row=row_idx, column=1).value
        if not rid_value:
            continue
        rid = str(rid_value).strip()
        if not re.match(r"^R\d{4}$", rid):
            continue

        per_variant: dict[str, RuleApplicability] = {}
        for variant_name, col in variant_cols.items():
            sev = _coerce_severity(
                ws_var.cell(row=row_idx, column=col).value,
                rule_id=rid,
                variant=variant_name,
            )
            if sev is None:
                continue
            add_raw = ws_var.cell(row=row_idx, column=col + 1).value
            per_variant[variant_name] = RuleApplicability(
                severity=sev,
                additional_values=_parse_additional_values(add_raw),
            )
        applicability[rid] = per_variant

    # ─── Requirements tab: titles, text, notes, discussion ─────────────────────
    ws_req = wb[_REQS_SHEET]
    requirements: list[Requirement] = []
    for row in ws_req.iter_rows(min_row=2, values_only=True):
        rid_v = row[0]
        if not rid_v:
            continue
        rid = str(rid_v).strip()
        if not re.match(r"^R\d{4}$", rid):
            continue

        text_id = (str(row[1]) if row[1] else "").strip()
        version = _coerce_version(row[2])
        title = (str(row[3]) if row[3] else "").strip()
        text = (str(row[4]) if row[4] else "").strip()
        notes = (str(row[5]) if row[5] else "").strip() or None
        discussion = (str(row[6]) if row[6] else "").strip() or None

        requirements.append(
            Requirement(
                id=rid,
                text_id=text_id,
                version=version,
                title=title,
                text=text,
                notes=notes,
                discussion=discussion,
                applicability=applicability.get(rid, {}),
            )
        )

    log.info("Parsed %d requirements", len(requirements))

    return RequirementManifest(
        parsed_at=datetime.now(UTC),
        parser_version=__version__,
        spec_xlsx_sha256=sha256_file(path),
        variants=variants,
        requirements=requirements,
    )
